import module as m
import module._class as c
import _dat
from environment import settings as sets

import os
import pandas as pd
import numpy as np
import lightgbm as lgb
import openpyxl as xl
import datetime
import re

from tabulate import tabulate
from openpyxl.styles import PatternFill


def predict(horse_results, peds, results, date=sets.date.replace('/', ''), venue_id_list=sets.venue_id_list):
    # dateからdayを取得
    if date[6] == '0':
        day = date[7]
    else:
        day = date[6:]
    # 過去のデータに対する予測を正確に行うために、訓練データから対象外のデータを削除する
    data_c = results.data_c[results.data_c['date'] < datetime.datetime(int(sets.year), int(sets.month), int(day))]
    # 説明変数と目的変数に分割
    X = data_c.drop(sets.drop_list+['odds', 'rank_binary', 'rank_regression', 'rank_lambdarank'], axis=1)
    y = data_c['rank_regression']
    X = X.drop('date', axis=1)

    # ランキング学習のためのクエリの作成
    query = list()
    for i in X.groupby(level=0):
        query.append(len(i[1]))
    # lightgbmでの学習
    lgb_clf = lgb.LGBMRanker(**sets.params)
    lgb_clf.fit(X.values, y.values, group=query)
    # 複勝の回収率の期待値が100%を超える基準が記されたdf
    fukusho_odds_df = m.get_fukusho_odds_df(lgb_clf, X, data_c)
    better_fukusho_df = pd.DataFrame([], columns=['race_id', 'horse_num', 'predict_order', 'odds', 'score'])
    # 出馬表データのスクレイピング
    # 欲しい出馬表のrace_id, 日付を引数とする。
    # 'https://db.netkeiba.com/race/' + race_id (データベース)
    # 'https://race.netkeiba.com/race/shutuba.html?race_id=' + race_id (出馬表)
    dir_path = '../results/'+sets.year+'/pdf/'+sets.month+'月/'
    for venue_id in venue_id_list:
        race_id_dict = {venue_id[4:6]: {venue_id[6:8]: [venue_id+str(i).zfill(2) for i in range(1, 13)]}}
        venue_name = [k for k, v in sets.place_dict.items() if v == venue_id[4:6]][0]
        sheet_name = day+'日'+venue_name
        file_path = '../results/'+sets.year+'/pdf/'+sets.month+'月/'+sheet_name+'.pdf'
        # 出馬表インスタンスの作成
        shutuba_table = c.ShutubaTable.scrape(race_id_dict, date, venue_id, results, horse_results, peds, n_samples=sets.n_samples)
        # ModelEvaluatorインスタンスの作成
        model_evaluator = c.ModelEvaluator(lgb_clf, sets.tables_path, kind=1, obj=sets.objective_type)
        X_fact = shutuba_table.data_c.drop(sets.drop_list+['date'], axis=1)
        # 各レースの本命馬、対抗馬、単穴馬、連下馬の出力
        pred = model_evaluator.predict_proba(X_fact, train=False)
        proba_table = shutuba_table.data_c[['horse_num']].astype('int').copy()
        proba_table['score'] = pred.astype('float64')
        # 払い戻し表のスクレイピング(まだサイトが完成していない場合はexceptに飛ぶ)
        try:
            exists_return_tables = 1
            today_return_tables, arrival_tables_df = c.Return.scrape(race_id_dict)
            rt = c.Return(today_return_tables)
            tansho_df = rt.tansho.fillna(0)
            fukusho_df = rt.fukusho.fillna(0)
            sanrentan_df = rt.sanrentan.fillna(0)
            sanrenpuku_df = rt.sanrenpuku.fillna(0)
            predict_df = pd.DataFrame(
                index=[],
                columns=sets.predict_columns+sets.result_columns
            )
            return1_df = pd.DataFrame(
                index=[],
                columns=sets.return1_columns
            )
            return2_df = pd.DataFrame(
                index=[],
                columns=sets.return2_columns
            )
        except KeyError:
            exists_return_tables = 0
            predict_df = pd.DataFrame(
                index=[],
                columns=sets.predict_columns
            )
        # 三連複が的中したかどうか記録する配列
        hits_sanrenpuku = [0] * shutuba_table.data_c.index.unique().shape[0]
        for i, proba in enumerate(proba_table.groupby(level=0)):
            # 使う頻度の高い数値を変数に格納
            race_proba = proba[1].sort_values('score', ascending=False).head(6)
            race_class = m.make_race_class(shutuba_table, proba)
            race_id = proba[0]
            if race_id[-2] == '0':
                race_num = ''+race_id[-1]
            else:
                race_num = race_id[-2:]
            today_odds_df = m.get_today_odds_df(race_id)
            _race_proba = race_proba.copy(deep=True)
            if len(today_odds_df):
                popular = today_odds_df.sort_values('tansho_odds').index
                today_df = m.determine_fukusho_odds(race_proba, today_odds_df, popular)
                today_df['score'] = proba[1].sort_values('horse_num').set_index('horse_num')
                # あるレースの買うべき複勝馬券を計算する
                for horse_num, row in today_df.iterrows():
                    if m.get_better_fukusho(fukusho_odds_df, row['score'], row['fukusho_odds']):
                        better_fukusho_df = better_fukusho_df.append(
                            pd.Series([
                                race_id,
                                horse_num,
                                list(proba[1].sort_values('score', ascending=False)['horse_num']).index(int(horse_num))+1,
                                row['fukusho_odds'],
                                row['score']
                            ], index=better_fukusho_df.columns),
                            ignore_index=True
                        )
                # 買うべき複勝馬に星マークをつける
                for _num in better_fukusho_df[better_fukusho_df['race_id'] == race_id]['horse_num']:
                    for i, _proba in enumerate(_race_proba.iterrows()):
                        if '☆' not in str(_proba[1]['horse_num']) and int(_proba[1]['horse_num']) == _num:
                            _race_proba.iat[i, 0] = '☆ '+str(_race_proba.iat[i, 0])
            # 本命馬のランクの決定
            if race_proba.iat[0, 1] - race_proba.iat[1, 1] >= 1 and race_proba.iat[0, 1] >= 2:
                fav_rank = 'A'
            elif race_proba.iat[0, 1] - race_proba.iat[1, 1] >= 1 or race_proba.iat[0, 1] >= 2:
                fav_rank = 'B'
            elif race_proba.iat[0, 1] - race_proba.iat[1, 1] >= 0.4:
                fav_rank = 'C'
            elif np.isnan(race_proba.iat[0, 1]) or (race_proba.iat[0, 1] == race_proba.iat[2, 1]):
                fav_rank = 'x'
                proba_table.loc[race_id, 'score'] = np.zeros(len(proba[1]))
            else:
                fav_rank = '-'
            # race_id_listの長さが6に満たない場合、'-'で埋める
            s = pd.Series(['-', '-'], index=race_proba.columns, name=race_proba.index[0])
            for j in range(6-len(race_proba)):
                race_proba = race_proba.append(s)
            # ここから先の処理は結果に対する処理(予測ではないため、結果が得られていないと計算できない)
            if exists_return_tables:
                # 本命馬の着順結果を取得
                real_arrival = list()
                for j in range(6):
                    if len(arrival_tables_df.loc[race_id, :][arrival_tables_df.loc[race_id, 'horse_num'] == race_proba.iat[j, 0]]['rank']) == 1:
                        real_arrival.append(
                            str(arrival_tables_df.loc[race_id, :][
                                    arrival_tables_df.loc[race_id, 'horse_num'] == race_proba.iat[j, 0]
                                ]['rank'].iat[0])
                            )
                    else:
                        real_arrival.append('-')
                # 単勝オッズを取得
                tansho_odds = tansho_df.loc[race_id, 'return'] / 100
                # 複勝オッズを取得
                fukusho_odds = [fukusho_df.loc[race_id, 'return_1'] / 100,
                                fukusho_df.loc[race_id, 'return_2'] / 100,
                                fukusho_df.loc[race_id, 'return_3'] / 100]
                # 単勝回収率の計算
                if real_arrival[0] == '1':
                    tansho_money = tansho_df.loc[race_id, 'return']
                else:
                    tansho_money = 0
                # 複勝回収率の計算
                # if '1' == real_arrival[0]: fukusho_money = fukusho_df.loc[race_id, 'return_0']
                # elif '2' == real_arrival[0]: fukusho_money = fukusho_df.loc[race_id, 'return_1']
                # elif '3' == real_arrival[0]: fukusho_money = fukusho_df.loc[race_id, 'return_2']
                fukusho_money = 0
                for i, _proba in enumerate(_race_proba.iterrows()):
                    if '☆' in str(_proba[1]['horse_num']) and real_arrival[i].isnumeric() and int(real_arrival[i]) <= 3:
                        fukusho_money = fukusho_df.loc[race_id, 'return_'+real_arrival[i]]
                # 三連単の結果を取得
                sanrentan_results = str(int(sanrentan_df.loc[race_id, 'win_0']))+' → ' + \
                    str(int(sanrentan_df.loc[race_id, 'win_1']))+' → ' + \
                    str(int(sanrentan_df.loc[race_id, 'win_2']))
                # 三連単のオッズを取得
                sanrentan_odds = sanrentan_df.loc[race_id, 'return'] / 100
                # 三連複のオッズを取得
                sanrenpuku_odds = sanrenpuku_df.loc[race_id, 'return'] / 100
                # 三連複が的中したかどうか記録
                sanrenpuku_results_list = list(map(int, sanrenpuku_df.loc[race_id, 'win_0':'win_2']))
                sanrenpuku_predict_list = [
                                            race_proba.iat[0, 0],
                                            race_proba.iat[1, 0],
                                            race_proba.iat[2, 0],
                                            race_proba.iat[3, 0],
                                            race_proba.iat[4, 0],
                                            race_proba.iat[5, 0]
                                        ]
                for result in sanrenpuku_results_list:
                    if result in sanrenpuku_predict_list:
                        hits_sanrenpuku[i] += 1
                # 三連複の回収率の計算
                if hits_sanrenpuku[i] == 3:
                    sanrenpuku_money = sanrenpuku_df.loc[race_id, 'return']
                else:
                    sanrenpuku_money = 0
                # 三連単の回収率の計算
                if hits_sanrenpuku[i] == 3 and real_arrival[0] == '1':
                    sanrentan_money = sanrentan_df.loc[race_id, 'return']
                else:
                    sanrentan_money = 0
                # データをdfに追加
                predict_df.loc[race_num+'R'] = [
                                                fav_rank,
                                                race_class,
                                                str(_race_proba.iat[0, 0])+' ('+real_arrival[0]+')',
                                                str(_race_proba.iat[1, 0])+' ('+real_arrival[1]+')',
                                                str(_race_proba.iat[2, 0])+' ('+real_arrival[2]+')',
                                                str(_race_proba.iat[3, 0])+' ('+real_arrival[3]+')',
                                                str(_race_proba.iat[4, 0])+' ('+real_arrival[4]+')',
                                                str(_race_proba.iat[5, 0])+' ('+real_arrival[5]+')',
                                                sanrentan_results,
                                                proba[1]['horse_num'][-1],
                                                tansho_odds,
                                                fukusho_odds[0],
                                                fukusho_odds[1],
                                                fukusho_odds[2],
                                                sanrentan_odds,
                                                sanrenpuku_odds,
                                                tansho_money,
                                                fukusho_money,
                                                sanrentan_money,
                                                sanrenpuku_money
                                                ]
            else:
                # データをdfに追加
                predict_df.loc[race_num+'R'] = [
                                                fav_rank,
                                                race_class,
                                                _race_proba.iat[0, 0],
                                                _race_proba.iat[1, 0],
                                                _race_proba.iat[2, 0],
                                                _race_proba.iat[3, 0],
                                                _race_proba.iat[4, 0],
                                                _race_proba.iat[5, 0],
                                                proba[1]['horse_num'][-1]
                                                ]

        if exists_return_tables:
            # predict_dfをexcelに追記
            decorate_excel = c.DecorateExcel(predict_df, sets.excel_path, sheet_name, mode='new')
            decorate_excel.decorate_excelsheet()
            decorate_excel.coloring_by_score(proba_table, venue_id)
            decorate_excel.save()

            # return1_dfの作成
            all_win = np.zeros(4, dtype=int)
            all_sanrenpuku_win = 0
            all_sanrentan_win = 0
            all_rank_cnt = 0
            all_sanren_cnt = 0
            all_tansho_money = 0
            all_fukusho_money = 0
            all_sanrentan_money = 0
            all_sanrenpuku_money = 0
            # 各ランク毎の結果を集計
            for rank in ['A', 'B', 'C', '-']:
                tansho_win = np.zeros(4, dtype=int)  # 勝ち馬を予測できた数
                rank_cnt = 0  # ランク毎の馬の数
                tansho_money = 0  # 単勝の回収率
                fukusho_money = 0  # 複勝回収率
                sanrenpuku_win = 0  # 三連複的中数
                sanren_cnt = 0  # 三連系を購入すると決定した数
                sanrenpuku_money = 0  # 三連複回収金額
                sanrentan_win = 0  # 三連単的中数
                sanrentan_money = 0  # 三連単回収金額
                wb = xl.load_workbook(sets.excel_path)
                ws = wb[sheet_name]
                for row in ws.rows:
                    for cell in row:
                        coord = cell.coordinate
                        if ws.cell(column=cell.column, row=decorate_excel.col_heading).value == '本命馬ランク' and ws[coord].value == rank:
                            rank_cnt += 1
                            for i in range(1, ws.max_column):
                                if ws.cell(column=i+1, row=decorate_excel.col_heading).value == '1着予想◎':
                                    if ws.cell(column=i+1, row=cell.row).value[-3:] in ['(1)', '(2)', '(3)']:
                                        for j in range(1, ws.max_column):
                                            if ws.cell(column=j+1, row=decorate_excel.col_heading).value == '複勝回収金額':
                                                fukusho_money += ws.cell(column=j+1, row=cell.row).value
                                    if ws.cell(column=i+1, row=cell.row).value[-3:] == '(1)':
                                        tansho_win[0] += 1
                                        for j in range(1, ws.max_column):
                                            if ws.cell(column=j+1, row=decorate_excel.col_heading).value == '単勝回収金額':
                                                tansho_money += ws.cell(column=j+1, row=cell.row).value
                                    elif ws.cell(column=i+1, row=cell.row).value[-3:] == '(2)':
                                        tansho_win[1] += 1
                                    elif ws.cell(column=i+1, row=cell.row).value[-3:] == '(3)':
                                        tansho_win[2] += 1
                                    else:
                                        tansho_win[3] += 1
                            sanren_cnt += 1
                            for i in range(1, ws.max_column):
                                if ws.cell(column=i+1, row=decorate_excel.col_heading).value == '三連複オッズ':
                                    # True: 三連複予測成功
                                    if ws[ws.cell(column=i+1, row=cell.row).coordinate].fill == PatternFill(patternType='solid', fgColor='ffbf7f'):
                                        sanrenpuku_win += 1
                                        for j in range(1, ws.max_column):
                                            if ws.cell(column=j+1, row=decorate_excel.col_heading).value == '三連複流し回収金額':
                                                sanrenpuku_money += ws.cell(column=j+1, row=cell.row).value
                                if ws.cell(column=i+1, row=decorate_excel.col_heading).value == '三連単オッズ':
                                    # True: 三連単予測成功
                                    if ws[ws.cell(column=i+1, row=cell.row).coordinate].fill == PatternFill(patternType='solid', fgColor='ffbf7f'):
                                        sanrentan_win += 1
                                        for j in range(1, ws.max_column):
                                            if ws.cell(column=j+1, row=decorate_excel.col_heading).value == '三連単流し回収金額':
                                                sanrentan_money += ws.cell(column=j+1, row=cell.row).value
                if rank != '-':
                    return1_df.loc[rank] = [
                                            str(tansho_win[0])+'-'+str(tansho_win[1])+'-'+str(tansho_win[2])+'-'+str(tansho_win[3]),
                                            str(tansho_win[0])+'/'+str(rank_cnt),
                                            str(tansho_win[0]+tansho_win[1]+tansho_win[2])+'/'+str(rank_cnt),
                                            str(sanrentan_win)+'/'+str(sanren_cnt),
                                            str(sanrenpuku_win)+'/'+str(sanren_cnt),
                                            m.div(tansho_money, 100*rank_cnt),
                                            m.div(fukusho_money, 100*rank_cnt),
                                            m.div(sanrentan_money, 2000*sanren_cnt),
                                            m.div(sanrenpuku_money, 2000*sanren_cnt)
                                        ]
                else:
                    return1_df.loc['ABC'] = [
                                                str(all_win[0])+'-'+str(all_win[1])+'-'+str(all_win[2])+'-'+str(all_win[3]),
                                                str(all_win[0])+'/'+str(all_rank_cnt),
                                                str(all_win[0]+all_win[1]+all_win[2])+'/'+str(all_rank_cnt),
                                                str(all_sanrentan_win)+'/'+str(all_sanren_cnt),
                                                str(all_sanrenpuku_win)+'/'+str(all_sanren_cnt),
                                                m.div(all_tansho_money, 100*all_rank_cnt),
                                                m.div(all_fukusho_money, 100*all_rank_cnt),
                                                m.div(all_sanrentan_money, 2000*all_sanren_cnt),
                                                m.div(all_sanrenpuku_money, 2000*all_sanren_cnt)
                                            ]
                all_win += tansho_win
                all_sanrentan_win += sanrentan_win
                all_sanrenpuku_win += sanrenpuku_win
                all_rank_cnt += rank_cnt
                all_sanren_cnt += sanren_cnt
                all_tansho_money += tansho_money
                all_fukusho_money += fukusho_money
                all_sanrentan_money += sanrentan_money
                all_sanrenpuku_money += sanrenpuku_money
                wb.save(sets.excel_path)
                wb.close()

            return1_df.loc['全体1'] = [
                                        str(all_win[0])+'-'+str(all_win[1])+'-'+str(all_win[2])+'-'+str(all_win[3]),
                                        str(all_win[0])+'/'+str(all_rank_cnt),
                                        str(all_win[0]+all_win[1]+all_win[2])+'/'+str(all_rank_cnt),
                                        str(all_sanrentan_win)+'/'+str(all_sanren_cnt),
                                        str(all_sanrenpuku_win)+'/'+str(all_sanren_cnt),
                                        m.div(all_tansho_money, 100*all_rank_cnt),
                                        m.div(all_fukusho_money, 100*all_rank_cnt),
                                        m.div(all_sanrentan_money, 2000*all_sanren_cnt),
                                        m.div(all_sanrenpuku_money, 2000*all_sanren_cnt)
                                    ]

            # return2_dfの作成
            fukusho_win = np.zeros(3, dtype=int)
            wide_win = 0
            sanrenpuku_win = 0
            sanrentan_win = 0
            no_cnt = 0
            wb = xl.load_workbook(sets.excel_path)
            ws = wb[sheet_name]
            for row in ws.rows:
                hits_sanrenpuku = 0
                winning_horse = np.zeros(3, dtype=int)
                for cell in row:
                    coord = cell.coordinate
                    if cell.row == decorate_excel.col_heading:
                        break
                    if ws.cell(column=cell.column, row=decorate_excel.col_heading).value == '本命馬ランク' and ws[coord].value == 'x':
                        no_cnt += 1
                        break
                    else:
                        for i, col_name in enumerate(['1着予想◎', '2着予想○', '3着予想▲']):
                            if ws.cell(column=cell.column, row=decorate_excel.col_heading).value == col_name:
                                try:
                                    winning_horse[i] = (lambda x: x if x < 4 else 0)(int(re.sub('[()]', '', ws[coord].value[-3:])))
                                except Exception:
                                    winning_horse[i] = 0
                                if winning_horse[i] > 0:
                                    fukusho_win[i] += 1
                if np.all([i in winning_horse for i in [1, 2]]):
                    wide_win += 1
                if np.count_nonzero(winning_horse) == 3:
                    sanrenpuku_win += 1
                if np.all(winning_horse == np.array([1, 2, 3])):
                    sanrentan_win += 1
            return2_df.loc['全体2'] = [
                                        str(wide_win)+'/'+str(12-no_cnt),
                                        str(sanrenpuku_win)+'/'+str(12-no_cnt),
                                        str(sanrentan_win)+'/'+str(12-no_cnt)
                                    ]
            wb.save(sets.excel_path)
            wb.close()

            # return1_dfをexcelに追記
            decorate_excel = c.DecorateExcel(return1_df, sets.excel_path, sheet_name)
            decorate_excel.decorate_excelsheet(
                                                pct_list=['単勝回収率', '複勝回収率', '三連単流し回収率', '三連複流し回収率']
                                                )
            decorate_excel.save()

            # return2_dfをexcelに追記
            decorate_excel = c.DecorateExcel(return2_df, sets.excel_path, sheet_name)
            decorate_excel.decorate_excelsheet()
            decorate_excel.save()

            # ディレクトリを自動作成
            if not os.path.isdir(dir_path):
                os.makedirs(dir_path)
                os.makedirs(dir_path.replace('pdf', 'png'))
            # xlsxをpdfに変換し、pdfディレクトリに保存する
            m.xlsx2pdf(file_path, ws)
            # pdfをpngに変換し、pngディレクトリに保存する
            m.pdf2png(file_path)

            # 現在時点での1ヶ月間の総収支をまとめたシートを作成
            m.make_BoP_sheet(sets.month+'月収支', sets.excel_path, file_path.replace(sheet_name, sets.month+'月収支'), sets.total_columns)
            m.make_hit_sheet(sets.month+'月的中率', sets.excel_path, file_path.replace(sheet_name, sets.month+'月的中率'), sets.total_columns)

            # 出力結果が得られた要因
            print(model_evaluator.feature_importance(X, type='split'))
        # (まだレースが開催されていない場合)予想結果の表示
        else:
            print('<'+venue_name+'>')
            print(tabulate(predict_df.iloc[:, :8], ['レース番号']+list(predict_df.columns), tablefmt='presto', showindex=True))
    print(better_fukusho_df.set_index(['race_id', 'horse_num']))


if __name__ == '__main__':
    # インスタンスの作成
    horse_results = c.HorseResults(_dat.horse_results['overall'])
    peds = c.Peds(_dat.ped_results['overall'])
    results = c.Results(_dat.race_results['overall'], horse_results, peds, n_samples=sets.n_samples)
    if sets.all_date:
        date_list = m.held_date_in_month(sets.session, sets.year+sets.month.zfill(2))
        for date in date_list:
            print(date)
            venue_id_list = sets.make_venue_id_list(date)
            print(venue_id_list)
            predict(horse_results, peds, results, date, venue_id_list)
    else:
        predict(horse_results, peds, results)
