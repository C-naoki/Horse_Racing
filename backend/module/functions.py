import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tqdm import tqdm

import openpyxl as xl
import re
import requests

from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import mm
from reportlab.lib import colors
from pathlib import Path
from pdf2image import convert_from_path
from bs4 import BeautifulSoup
from selenium import webdriver


def update_data(old, new):
    """
    Parameters:
    ----------
    old : pandas.DataFrame
        古いデータ
    new : pandas.DataFrame
        新しいデータ
    """

    filtered_old = old[~old.index.isin(new.index)]
    return pd.concat([filtered_old, new])


def all_update_data(dict, year_list=['2017', '2018', '2019', '2020', '2021', '2022']):
    ans = pd.DataFrame()
    for year in year_list:
        ans = update_data(ans, dict[year])
    return ans


def split_data(df, test_size=0.3):
    df_i = df.reset_index()
    sorted_id_list = df_i.sort_values(['date', 'index']).index.unique()
    train_id_list = sorted_id_list[: round(len(sorted_id_list) * (1 - test_size))]
    test_id_list = sorted_id_list[round(len(sorted_id_list) * (1 - test_size)):]
    train = df_i.loc[train_id_list]
    test = df_i.loc[test_id_list]
    return train.set_index('index'), test.set_index('index')


def gain(return_func, X, n_samples=100, t_range=[0.5, 3.5]):
    gain = {}
    for i in tqdm(range(n_samples)):
        # min_thresholdから1まで、n_samples等分して、thresholdをfor分で回す
        threshold = t_range[1] * i / n_samples + t_range[0] * (1-(i/n_samples))
        n_bets, return_rate, n_hits, std = return_func(X, threshold)
        if n_bets > 2:
            gain[threshold] = {
                'return_rate': return_rate,
                'n_hits': n_hits,
                'std': std,
                'n_bets': n_bets
            }
        if n_bets == 0:
            break
    return pd.DataFrame(gain).T


def plot(df, label=' '):
    # 標準偏差で幅をつけて薄くプロット
    plt.fill_between(
        df.index,
        y1=df['return_rate']-df['std'],
        y2=df['return_rate']+df['std'],
        alpha=0.3
    )  # alphaで透明度を設定
    # 回収率を実線でプロット
    plt.plot(df.index, df['return_rate'], label=label)
    plt.legend()  # labelで設定した凡例を表示させる
    plt.grid(True)  # グリッドをつける


def xlsx2pdf(pdf_file, ws, pagesize=[600, 250]):
    doc = SimpleDocTemplate(pdf_file, pagesize=(pagesize[0]*mm, pagesize[1]*mm))
    pdfmetrics.registerFont(TTFont('meiryo', '/Users/naoki/Downloads/font/meiryo/meiryo.ttc'))
    pdf_data = []
    data = []
    # Tableの作成
    for row in ws.rows:
        row_list = []
        for cell in row:
            if cell.number_format == '0.00%' and cell.value != '-':
                row_list.append(str(round(cell.value*100, 2))+'%')
            else:
                row_list.append(cell.value)
        data.append(row_list)
    tt = Table(data)
    tt.setStyle(TableStyle([
                                ('FONT', (0, 0), (-1, -1), 'meiryo', 11),
                                ('GRID', (0, 0), (ws.max_column, ws.max_row), 0.25, colors.black),
                                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER')
                            ]))
    # 着色
    for row in ws.rows:
        for cell in row:
            cell_idx = (column_index_from_string(cell.coordinate[0])-1, int(cell.coordinate[1:])-1)
            back_color_code = ws[cell.coordinate].fill.fgColor.rgb[2:]
            text_color_code = ws[cell.coordinate].font.color.rgb[2:]
            back_color_name = code2name(back_color_code)
            text_color_name = code2name(text_color_code)
            tt.setStyle(TableStyle([
                                    ('BACKGROUND', cell_idx, cell_idx, back_color_name),
                                    ('TEXTCOLOR', cell_idx, cell_idx, text_color_name)
                                    ]))
    pdf_data.append(tt)
    doc.build(pdf_data)


def code2name(color_code):
    name_code_dict = {
        'ffbf7f': colors.lightsalmon,
        'a8d3ff': colors.lightblue,
        'd3d3d3': colors.lightgrey,
        '50C878': colors.lightgreen,
        '000000': colors.black,
        'ffffff': colors.white
    }
    if color_code in name_code_dict:
        return name_code_dict[color_code]
    else:
        return colors.white


def pdf2png(pdf_path, dpi=200, fmt='png'):
    png_path = pdf_path.replace('pdf', 'png')
    pdf_path = Path(pdf_path)
    png_path = Path(png_path)
    page = convert_from_path(pdf_path, dpi)
    page[0].save(png_path, fmt)


def div(a, b, digit=4):
    if b == 0:
        return '-'
    else:
        return round(a / b, digit)


def output_return_rate(ws, total):
    if total[-3:] == 'ABC':
        rank = 'ABC'
    elif total[-1] not in ['A', 'B', 'C']:
        rank = '全体1'
    else:
        rank = total[-1]
    total = total.replace(rank, '')
    for col in ws.columns:
        for cell in col:
            if cell.value == total+'回収率':
                col_i = cell.column
            if cell.value == rank and cell.column == 1:
                row_i = cell.row
    return ws.cell(row=row_i, column=col_i).value


def output_hit_rate(ws, total):
    if total[-3:] == 'ABC':
        rank = 'ABC'
    elif total[-1] not in ['A', 'B', 'C']:
        if total not in ['ワイド', '三連複', '三連単']:
            rank = '全体1'
        else:
            rank = '全体2'
    else:
        rank = total[-1]
    total = total.replace(rank, '')
    for col in ws.columns:
        for cell in col:
            if cell.value == total+'的中率' or cell.value == total:
                col_i = cell.column
            if cell.value == rank and cell.column == 1:
                row_i = cell.row
    return ws.cell(row=row_i, column=col_i).value


# BoP: Balance of Payments
def make_BoP_sheet(total, excel_path, pdf_path, total_columns):
    wb = xl.load_workbook(excel_path)
    df = pd.DataFrame(columns=total_columns)
    sheetnames = list(wb.sheetnames)
    tansho_all_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}
    sanren_all_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}
    fukusho_all_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}
    all_rank_dict = {'単': tansho_all_rank, '三': sanren_all_rank, '複': fukusho_all_rank}
    sheetnames = [i for i in sheetnames if i not in [total, total[0]+'月的中率', 'Sheet1']]
    # 回収金額の計算
    for name in sheetnames:
        ws_ = wb[name]
        # rankとvenue_idごとのbet数
        tansho_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}
        sanren_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}
        fukusho_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}
        for i in range(2, 14):
            tansho_rank[ws_.cell(row=i, column=2).value] += 100
            sanren_rank[ws_.cell(row=i, column=2).value] += 2000
            fukusho_rank[ws_.cell(row=i, column=2).value] += 100
        rank_dict = {'単': tansho_rank, '三': sanren_rank, '複': fukusho_rank}
        for key in tansho_rank.keys():
            for kind in ['単', '三', '複']:
                all_rank_dict[kind][key] += rank_dict[kind][key]
        s = pd.Series([], name=name, index=[])
        for col in total_columns:
            s.loc[col] = output_return_rate(ws_, col)
        s.replace('-', 0, inplace=True)
        s = s-1
        for col in total_columns:
            for kind in ['単', '三', '複']:
                if col[0] == kind and 'ABC' in col:
                    s[col] *= sum(rank_dict[kind].values()) - rank_dict[kind]['-'] - rank_dict[kind]['x']
                elif col[0] == kind and col[-1] in ['A', 'B', 'C']:
                    s[col] *= rank_dict[kind][col[-1]]
                elif col[0] == kind and col[-1] not in ['A', 'B', 'C']:
                    s[col] *= sum(rank_dict[kind].values()) - rank_dict[kind]['x']
        df = df.append(s)
    total_s = pd.Series(df.sum(), name='合計(円)')
    total_s2 = pd.Series(index=total_columns, name='合計(%)', dtype=float)
    # 回収率の計算
    for col in total_columns:
        for kind in ['単', '三', '複']:
            if col[0] == kind and 'ABC' in col:
                total_s2[col] = div(
                    total_s[col]+sum(all_rank_dict[kind].values()) - all_rank_dict[kind]['-'] - all_rank_dict[kind]['x'],
                    sum(all_rank_dict[kind].values()) - all_rank_dict[kind]['-'] - all_rank_dict[kind]['x']
                )
            elif col[0] == kind and col[-1] in ['A', 'B', 'C']:
                total_s2[col] = div(
                    total_s[col]+all_rank_dict[kind][col[-1]],
                    all_rank_dict[kind][col[-1]]
                )
            elif col[0] == kind and col[-1] not in ['A', 'B', 'C']:
                total_s2[col] = div(
                    total_s[col]+sum(all_rank_dict[kind].values()) - all_rank_dict[kind]['x'],
                    sum(all_rank_dict[kind].values()) - all_rank_dict[kind]['x']
                )
    df = df.append(total_s)
    df = df.round(-1)
    df = df.append(total_s2)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    writer.book = wb
    writer.sheets = {ws_.title: ws_ for ws_ in wb.worksheets}
    df.to_excel(writer, sheet_name=total)
    writer.save()
    # 必要箇所にカラーリング
    ws = wb[total]
    side1 = Side(style='thin', color='000000')
    border1 = Border(top=side1, bottom=side1, left=side1, right=side1)
    for col in ws.columns:
        for i, cell in enumerate(col):
            coord = cell.coordinate
            ws[coord].border = border1
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=False
            )
            if coord[0] == 'A' or coord[1:] == '1':
                ws[coord].fill = PatternFill(patternType='solid', fgColor='000000')
                ws[coord].font = Font(color='ffffff')
            else:
                ws[coord].fill = PatternFill(patternType='solid', fgColor='ffffff')
                ws[coord].font = Font(color='000000')
            try:
                if coord[1:] != str(ws.max_row):
                    if cell.value > 0:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='ffbf7f')
                    elif cell.value < 0:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='a8d3ff')
                else:
                    if cell.value > 1:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='ffbf7f')
                    elif cell.value < 1:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='a8d3ff')
                    cell.number_format = '0.00%'
            except Exception:
                continue
        ws.column_dimensions[col[0].column_letter].width = 15
    # 月収支シートをファイルの先頭に移動させる
    for ws_ in wb.worksheets:
        i = i+1
        if total in ws.title:
            # 先頭までの移動シート枚数
            Top_Sheet = i - 1
            # シート移動
            wb.move_sheet(ws, offset=-Top_Sheet)
    # xlsxをpdfに変換し、pdfディレクトリに保存する
    xlsx2pdf(pdf_path, ws, pagesize=[500, 250])
    # pdfをpngに変換し、pngディレクトリに保存する
    pdf2png(pdf_path)
    wb.save(excel_path)
    wb.close()


def make_hit_sheet(total, excel_path, pdf_path, total_columns):
    wb = xl.load_workbook(excel_path)
    total_columns = total_columns+['ワイド', '三連複', '三連単']
    df = pd.DataFrame(columns=total_columns)
    sheetnames = list(wb.sheetnames)
    sheetnames = [i for i in sheetnames if i not in [total, total[0]+'月収支', 'Sheet1']]
    # 的中率の計算
    for name in sheetnames:
        ws_ = wb[name]
        s = pd.Series([], name=name, index=[])
        for col in total_columns:
            s.loc[col] = output_hit_rate(ws_, col)
        s.replace('0/0', np.nan, inplace=True)
        df = df.append(s)
    total_s = hit_sum(df, total_columns)
    df = df.append(total_s)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    writer.book = wb
    writer.sheets = {ws_.title: ws_ for ws_ in wb.worksheets}
    df.to_excel(writer, sheet_name=total)
    writer.save()
    # 必要箇所にカラーリング
    ws = wb[total]
    side1 = Side(style='thin', color='000000')
    border1 = Border(top=side1, bottom=side1, left=side1, right=side1)
    for col in ws.columns:
        for i, cell in enumerate(col):
            coord = cell.coordinate
            ws[coord].border = border1
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=False
            )
            if coord[0] == 'A' or coord[1:] == '1':
                ws[coord].fill = PatternFill(patternType='solid', fgColor='000000')
                ws[coord].font = Font(color='ffffff')
            else:
                ws[coord].fill = PatternFill(patternType='solid', fgColor='ffffff')
                ws[coord].font = Font(color='000000')
            try:
                if coord[1:] != str(ws.max_row):
                    if cell.value > 0.5:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='ffbf7f')
                    elif cell.value < 0.5:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='a8d3ff')
                else:
                    if cell.value > 0.5:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='ffbf7f')
                    elif cell.value < 0.5:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='a8d3ff')
                    cell.number_format = '0.00%'
            except Exception:
                continue
        ws.column_dimensions[col[0].column_letter].width = 15
    # 月的中率シートをファイルの先頭に移動させる
    for ws_ in wb.worksheets:
        i = i+1
        if total in ws.title:

            # 先頭までの移動シート枚数
            Top_Sheet = i - 1
            # シート移動
            wb.move_sheet(ws, offset=-Top_Sheet)
    # xlsxをpdfに変換し、pdfディレクトリに保存する
    xlsx2pdf(pdf_path, ws, pagesize=[900, 250])
    # pdfをpngに変換し、pngディレクトリに保存する
    pdf2png(pdf_path)
    wb.save(excel_path)
    wb.close()


def hit_sum(df, total_columns):
    total_s = pd.Series(index=total_columns, name='合計(%)', dtype=float)
    for col in total_columns:
        total = df.loc[:, col]
        all_top = 0
        all_bottom = 0
        for a_hit in total:
            if isinstance(a_hit, str):
                top, bottom = [int(n) for n in a_hit.split('/')]
            else:
                top, bottom = (0, 0)
            all_top += top
            all_bottom += bottom
            total_s[col] = div(all_top, all_bottom)
    return total_s


def isExist_date(session, date):
    url = 'https://db.netkeiba.com/race/list/' + date
    html = session.get(url)
    html.encoding = "EUC-JP"
    soup = BeautifulSoup(html.content, "html.parser")
    # レースのある日付だけを引っ張ってくる
    for true_date in soup.find_all('table')[0].find_all('td'):
        try:
            if re.findall(r'\d+', true_date.find('a').get('href'))[0] == date:
                return True
        except Exception:
            pass
    return False


def held_date_in_month(session, month):
    url = 'https://db.netkeiba.com/race/list/' + month
    html = session.get(url)
    html.encoding = "EUC-JP"
    soup = BeautifulSoup(html.content, "html.parser")
    date_list = list()
    for date in soup.find_all('table')[0].find_all('td'):
        try:
            date_list.append(re.findall(r'\d+', date.find('a').get('href'))[0])
        except Exception:
            pass
    return date_list


def make_venue_id_list(date, num='all'):
    venue_id_list = set()
    url = 'https://race.netkeiba.com/top/race_list_sub.html?kaisai_date=' + date
    r = requests.get(url)
    soup = BeautifulSoup(r.content, "html.parser")
    for race_list in soup.find_all('li'):
        venue_id_list.add(re.findall(r'\d+', race_list.find_all('a')[0].get('href'))[0][:10])
    if num == 'all':
        return list(venue_id_list)
    else:
        return list(venue_id_list)[:num]


def make_race_class(shutuba_table, proba):
    df = shutuba_table.data_c.loc[proba[0], ['class_未勝利', 'class_１勝クラス', 'class_２勝クラス', 'class_３勝クラス', 'class_オープン']]
    for col in df.columns:
        if sum(df[col]):
            return col.strip('class_')
    return '新馬'


def get_fukusho_odds_df(model, X, y, range=[0.1, 2.5], width=0.05):
    fukusho_df = pd.DataFrame({'score': model.predict(X), 'correct': y['rank_binary']})
    fukusho_odds_df = pd.DataFrame([], columns=['num', 'rate', 'odds'])
    for threshold in np.arange(range[0]-width, range[1]+width, width):
        threshold = round(threshold, len(str(int(1/width))))
        temp = fukusho_df.copy()
        # 指定範囲に属するデータの絞り込み
        if threshold != range[0]-width:
            temp = temp[threshold < temp['score']]
        if threshold != range[1]:
            temp = temp[temp['score'] < threshold+width]
        # index名の指定
        if threshold == range[0]-width:
            row_name = ' ~ {}'.format(range[0])
        elif threshold == range[1]:
            row_name = '{} ~ '.format(range[1])
        else:
            row_name = '{} ~ {}'.format(threshold, round(threshold+width, len(str(int(1/width)))))
        # dfに格納
        if len(temp[temp['correct'] == 1]):
            fukusho_odds_df.loc[row_name] = [len(temp), len(temp[temp['correct'] == 1])/len(temp), len(temp)/len(temp[temp['correct'] == 1])]
        else:
            fukusho_odds_df.loc[row_name] = [len(temp), len(temp[temp['correct'] == 1])/len(temp), 9999]
    return fukusho_odds_df


def get_better_fukusho(fukusho_odds_df, score, odd):
    index_list = fukusho_odds_df.index
    for index in index_list:
        hoge = index.split(' ~ ')
        try:
            if score < float(hoge[1]):
                std_odd = fukusho_odds_df.loc[index, 'odds']
                break
        except Exception:
            std_odd = fukusho_odds_df.iloc[-1]['odds']
    return std_odd <= odd


# seleniumによるスクレイピングの準備
def set_selenium() -> webdriver.chrome.webdriver.WebDriver:
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    driver = webdriver.Chrome(executable_path='/Users/naoki/programing/chromedriver', options=options)
    driver.implicitly_wait(15)
    return driver


def get_today_odds_df(race_id: str) -> pd.DataFrame:
    ans_df = pd.DataFrame()
    driver = set_selenium()
    driver.get('https://race.netkeiba.com/odds/index.html?type=b1&race_id='+race_id+'&rf=shutuba_submenu')
    html = driver.page_source.encode('utf-8')
    dfs = pd.read_html(str(BeautifulSoup(html, "html.parser").html))
    try:
        ans_df['horse_num'] = dfs[0]['馬番']
        ans_df['tansho_odds'] = dfs[0]['オッズ']
        ans_df['fukusho_odds'] = dfs[1]['オッズ']
        ans_df.set_index('horse_num', inplace=True)
        ans_df['fukusho_odds'] = ans_df.loc[:, 'fukusho_odds']
        return ans_df
    except KeyError:
        return pd.DataFrame()


def determine_fukusho_odds(race_proba, today_odds_df, popular):
    df = today_odds_df.copy()
    hoge = set(list(popular[:3]) + list(race_proba['horse_num'][:3]))
    df = df[df['tansho_odds'] != '---.-']
    df['fukusho_odds'] = df.loc[:, 'fukusho_odds'].map(
        lambda x: round(float(x.split(' - ')[0]) + (len(hoge)-3)*(float(x.split(' - ')[1]) - float(x.split(' - ')[0]))/9, 2))
    return df.astype('float64')
