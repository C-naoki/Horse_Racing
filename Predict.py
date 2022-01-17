import module as m
import module._class as c
import _dat

import pandas as pd
import numpy as np
import optuna.integration.lightgbm as lgb_o
import lightgbm as lgb
import openpyxl as xl
import datetime

from environment.variables import *
from tabulate import tabulate
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment

if __name__ == '__main__':
    venue_id_list = ["2022060106", "2022070106", "2022100102"] # race_id_list
    date = '2022/01/16' # レース日
    year = date[0:4]
    month = date[5:7]
    day = date[8:10]
    dt_now = datetime.datetime.now() # 今の日付
    race_date = datetime.date(int(year), int(month), int(day)) # レース日
    today = datetime.date(dt_now.year, dt_now.month, dt_now.day)
    if month[0] == '0': month = month[1]
    if day[0] == '0': day = day[1]
    excel_path = 'results/'+year+'/xlsx/'+month+'月.xlsx'
    # race_dataの取得
    r = c.Results(_dat.race_results)
    # 前処理
    r.preprocessing()
    # 馬の過去成績の追加
    hr = c.HorseResults(_dat.horse_results)
    r.merge_horse_results(hr)
    # 5世代分の血統データの追加
    p = c.Peds(_dat.ped_results)
    p.encode()
    r.merge_peds(p.peds_e)
    # カテゴリ変数の処理
    r.process_categorical()
    print("\n<finish making race results>\n")

    X = r.data_c.drop(['rank', 'date', '単勝'], axis=1)
    y = r.data_c['rank']

    params={'objective': 'binary',
            'random_state': 100,
            'feature_pre_filter': False,
            'lambda_l1': 5.656720411334836,
            'lambda_l2': 0.0008932155945833474,
            'num_leaves': 31,
            'feature_fraction': 0.4,
            'bagging_fraction': 1.0,
            'bagging_freq': 0,
            'min_child_samples': 20}
    lgb_clf = lgb.LGBMClassifier(**params)
    lgb_clf.fit(X.values, y.values)

    # 出馬表データのスクレイピング
    # 欲しい出馬表のrace_id, 日付を引数とする。
    # "https://db.netkeiba.com/race/" + race_id (データベース)
    # "https://race.netkeiba.com/race/shutuba.html?race_id=" + race_id (出馬表)
    for venue in venue_id_list:
        race_id_list = [venue + str(i).zfill(2) for i in range(1, 13)]
        # 出馬表のスクレイピング
        st = c.ShutubaTable.scrape(race_id_list=race_id_list, date=date)
        st.preprocessing()
        st.merge_horse_results(hr)
        st.merge_peds(p.peds_e)
        st.process_categorical(r.le_horse, r.le_jockey, r.data_pe)
        print("\n<finish making race card>\n")
        # ModelEvaluator
        me = c.ModelEvaluator(lgb_clf, ['_dat/pickle/overall/return_tables.pickle'])
        X_fact = st.data_c.drop(['date'], axis=1)
        # 各レースの本命馬、対抗馬、単穴馬、連下馬の出力
        venue_name = [k for k, v in place_dict.items() if v == race_id_list[0][4:6]][0]
        sheet_name = day+"日"+venue_name
        pred = me.predict_proba(X_fact, train=False)
        proba_table = st.data_c[['馬番']].astype('int').copy()
        proba_table['score'] = pred.astype('float64')
        print("\n~Expected results~")
        pd.options.display.float_format = '{:.4f}'.format
        # 払い戻し表のスクレイピング(まだサイトが完成していない場合はexceptに飛ぶ)
        try:
            return_chk = 1
            today_return_tables, arrival_tables_df = c.Return.scrape(race_id_list)
            rt = c.Return(today_return_tables)
            tansho_df = rt.tansho
            fukusho_df = rt.fukusho
            sanrenpuku_df = rt.sanrenpuku
            predict_df = pd.DataFrame(
                                        index = [], 
                                        columns = ["本命馬ランク", "三連複ランク", "本命馬◎", "対抗馬○", "単穴馬▲", "連下馬1△", "連下馬2△", "連下馬3△", "本命馬着順", "単勝オッズ", "三連複結果", "三連複オッズ", "単勝回収金額", "三連複回収金額"]
                                    )
        except:
            return_chk = 0
            predict_df = pd.DataFrame(
                                    index = [], 
                                    columns = ["本命馬ランク", "三連複ランク", "本命馬◎", "対抗馬○", "単穴馬▲", "連下馬1△", "連下馬2△", "連下馬3△"]
                                )
        # 三連複が的中したかどうか記録する配列
        sanrenpuku_chk = [0] * len(set(st.data_c.index))
        for i, proba in enumerate(proba_table.groupby(level=0)):
            if proba[0][-2] == "0":
                race_num = " "+proba[0][-1]
            else:
                race_num = proba[0][-2:]
            race_proba = proba[1].sort_values('score', ascending = False).head(6)
            # 三連複のランクの決定
            if race_proba.iat[min(5, len(race_proba)-1), 1] < 0:
                triple_rank = "C"
            elif race_proba.iat[min(5, len(race_proba)-1), 1] < 0.5:
                triple_rank = "B"
            else:
                triple_rank = "A"
            # 本命馬のランクの決定
            if race_proba.iat[0, 1] - race_proba.iat[1, 1] < 0.4:
                favorite_rank = "C"
            elif race_proba.iat[0, 1] - race_proba.iat[1, 1] < 1:
                favorite_rank = "B"
            else:
                favorite_rank = "A"
            # race_id_listの長さが6に満たない場合、"-"で埋める
            s = pd.Series(['-', '-'], index=race_proba.columns, name=race_proba.index[0])
            for j in range(6-len(race_proba)):
                race_proba=race_proba.append(s)
            if return_chk:
                # 本命馬の着順結果を取得
                real_arrival = arrival_tables_df.loc[proba[0],:][arrival_tables_df.loc[proba[0],"馬番"]==race_proba.iat[0, 0]]["着順"].iat[0]
                if real_arrival != "除": real_arrival = int(real_arrival)
                # 単勝オッズを取得
                tansho_odds = tansho_df.loc[proba[0], "return"] / 100
                # 単勝回収率の計算
                if real_arrival == 1: tansho_money = tansho_df.loc[proba[0], "return"]
                else: tansho_money = 0
                # 三連複の結果を取得
                sanrenpuku_results = str(sanrenpuku_df.loc[proba[0], "win_0"])+" - "+str(sanrenpuku_df.loc[proba[0], "win_1"])+" - "+str(sanrenpuku_df.loc[proba[0], "win_2"])
                # 三連複のオッズを取得
                sanrenpuku_odds = sanrenpuku_df.loc[proba[0], "return"] / 100
                # 三連複が的中したかどうか記録
                sanrenpuku_results_list = list(map(int, sanrenpuku_df.loc[proba[0], "win_0":"win_2"]))
                sanrenpuku_predict_list = [race_proba.iat[0, 0], race_proba.iat[1, 0], race_proba.iat[2, 0], race_proba.iat[3, 0], race_proba.iat[4, 0], race_proba.iat[5, 0]]
                for result in sanrenpuku_results_list:
                    if result in sanrenpuku_predict_list:
                        sanrenpuku_chk[i] += 1
                # 三連複の回収率の計算
                if sanrenpuku_chk[i] == 3: sanrenpuku_money = sanrenpuku_df.loc[proba[0], "return"]
                else: sanrenpuku_money = 0
                # データをdfに追加
                predict_df.loc[race_num+"R"] = [favorite_rank, triple_rank, race_proba.iat[0, 0], race_proba.iat[1, 0], race_proba.iat[2, 0], race_proba.iat[3, 0], race_proba.iat[4, 0], race_proba.iat[5, 0], real_arrival, tansho_odds, sanrenpuku_results, sanrenpuku_odds, tansho_money, sanrenpuku_money]
            else:
                # データをdfに追加
                predict_df.loc[race_num+"R"] = [favorite_rank, triple_rank, race_proba.iat[0, 0], race_proba.iat[1, 0], race_proba.iat[2, 0], race_proba.iat[3, 0], race_proba.iat[4, 0], race_proba.iat[5, 0]]
        sanrenpuku_chk = [0, *sanrenpuku_chk]
        # (まだレースが開催されていない場合)xlsxファイルに保存するデータの表示
        if return_chk:
            # データをxlsxファイルに書き込む
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode="a", if_sheet_exists="replace") as writer:
                predict_df.to_excel(writer, sheet_name=day+"日"+venue_name)
            # 追加したデータの修正
            wb = xl.load_workbook(excel_path)
            ws = wb[sheet_name]
            # 表の左上に開催地の追加
            ws['A1'] = venue_name
            ws['A1'].font = Font(bold=True)
            side1 = Side(style='thin', color='000000')
            side2 = Side(style='double', color='000000')
            border1 = Border(top=side1, bottom=side1, left=side1, right=side1)
            border2 = Border(top=side1, bottom=side1, left=side1, right=side2)
            for col in ws.columns:
                max_length = 0
                for i, cell in enumerate(col):
                    max_length = max(max_length, len(str(cell.value)))
                    if cell.coordinate[0] == 'I': ws[cell.coordinate].border = border2
                    else: ws[cell.coordinate].border = border1
                    # ランクA及び着順の予想が的中した時、セルをオレンジ色にする
                    if ws[cell.coordinate].value == 'A' or (cell.coordinate[0] == 'J' and cell.value == 1) or (cell.coordinate[0] == 'L' and sanrenpuku_chk[i] == 3):
                        ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor='ffa500')
                    # ランクB及び着順の予想が3着以内及び三連複の3頭中2頭的中した時、セルを水色にする
                    if ws[cell.coordinate].value == 'B' or (cell.coordinate[0] == 'J' and (cell.value == 2 or cell.value == 3)) or (cell.coordinate[0] == 'L' and sanrenpuku_chk[i] == 2):
                        ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor='87ceeb')
                    # ランクC及び及び三連複の3頭中1頭のみ的中した時、セルを水色にする
                    if ws[cell.coordinate].value == 'C' or (cell.coordinate[0] == 'L' and sanrenpuku_chk[i] == 1):
                        ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor='d3d3d3')
                    if cell.coordinate[0] == 'A' or cell.coordinate[1:] == '1':
                        ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor='000000')
                        ws[cell.coordinate].font = Font(color="ffffff")
                    cell.alignment = Alignment(horizontal = 'center', 
                                        vertical = 'center',
                                        wrap_text = False)
                adjusted_width = max_length * 2.08
                ws.column_dimensions[col[0].column_letter].width = adjusted_width
            wb.save(excel_path)
        else:
            print("<"+venue_name+">")
            print(tabulate(predict_df, predict_df.columns, tablefmt="presto", showindex=True))
    # 出力結果が得られた要因
    print(me.feature_importance(X))