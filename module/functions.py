import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tqdm import tqdm

import lightgbm as lgb
import openpyxl as xl
import re
import requests

from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import mm
from reportlab.lib import colors
from pathlib import Path
from pdf2image import convert_from_path
from bs4 import BeautifulSoup

def update_data(old, new):
    """
    Parameters:
    ----------
    old : pandas.DataFrame
        古いデータ
    new : pandas.DataFrame
        新しいデータ
    """

    filtered_old = old[~old.index.isin(new.index)]
    return pd.concat([filtered_old, new])

def all_update_data(dict, year_list=['2017', '2018', '2019', '2020', '2021', '2022']):
    ans = pd.DataFrame()
    for year in year_list:
        ans = update_data(ans, dict[year])
    return ans

def split_data(df, test_size=0.3):
    sorted_id_list = df.sort_values('date').index.unique()
    train_id_list = sorted_id_list[: round(len(sorted_id_list) * (1 - test_size))]
    test_id_list = sorted_id_list[round(len(sorted_id_list) * (1 - test_size)) :]
    train = df.loc[train_id_list]
    test = df.loc[test_id_list]
    return train, test

def gain(return_func, X, n_samples=100, t_range=[0.5, 3.5]):
    gain = {}
    for i in tqdm(range(n_samples)):
        #min_thresholdから1まで、n_samples等分して、thresholdをfor分で回す
        threshold = t_range[1] * i / n_samples + t_range[0] * (1-(i/n_samples))
        n_bets, return_rate, n_hits, std = return_func(X, threshold)
        if n_bets > 2:
            gain[threshold] = {'return_rate': return_rate, 
                            'n_hits': n_hits,
                            'std': std,
                            'n_bets': n_bets}
        if n_bets == 0:
            break
    return pd.DataFrame(gain).T

def plot(df, label=' '):
    #標準偏差で幅をつけて薄くプロット
    plt.fill_between(df.index, y1=df['return_rate']-df['std'],
                y2=df['return_rate']+df['std'],
                alpha=0.3) #alphaで透明度を設定
    #回収率を実線でプロット
    plt.plot(df.index, df['return_rate'], label=label)
    plt.legend() #labelで設定した凡例を表示させる
    plt.grid(True) #グリッドをつける

def xlsx2pdf(pdf_file, ws, pagesize=[500, 250]):
    doc = SimpleDocTemplate( pdf_file, pagesize=(pagesize[0]*mm, pagesize[1]*mm) )
    pdfmetrics.registerFont(TTFont('meiryo', '/Users/naoki/Downloads/font/meiryo/meiryo.ttc'))
    pdf_data = []
    data = []
    # Tableの作成
    for row in ws.rows:
        row_list = []
        for cell in row:
            if cell.number_format == '0.00%' and cell.value != '-': row_list.append(str(round(cell.value*100, 2))+'%')
            else: row_list.append(cell.value)
        data.append(row_list)
    tt = Table(data)
    tt.setStyle(TableStyle([
                                ('FONT', (0, 0), (-1, -1), 'meiryo', 11),
                                ('GRID', (0, 0), (ws.max_column, ws.max_row), 0.25, colors.black),
                                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER')
                            ]))
    # 着色
    for row in ws.rows:
        for cell in row:
            cell_idx = (column_index_from_string(cell.coordinate[0])-1, int(cell.coordinate[1:])-1)
            back_color_code = ws[cell.coordinate].fill.fgColor.rgb[2:]
            text_color_code = ws[cell.coordinate].font.color.rgb[2:]
            back_color_name = code2name(back_color_code)
            text_color_name = code2name(text_color_code)
            tt.setStyle(TableStyle([
                                    ('BACKGROUND' ,cell_idx, cell_idx, back_color_name),
                                    ('TEXTCOLOR'  ,cell_idx, cell_idx, text_color_name)
                                    ]))
    pdf_data.append(tt)
    doc.build(pdf_data)

def code2name(color_code):
    name_code_dict = {
        'ffbf7f': colors.lightsalmon,
        'a8d3ff': colors.lightblue,
        'd3d3d3': colors.lightgrey,
        '50C878': colors.lightgreen,
        '000000': colors.black,
        'ffffff': colors.white
    }
    if color_code in name_code_dict: return name_code_dict[color_code]
    else: return colors.white

def pdf2png(pdf_path, dpi=200, fmt='png'):
    png_path = pdf_path.replace('pdf', 'png')
    pdf_path = Path(pdf_path)
    png_path = Path(png_path)
    page = convert_from_path(pdf_path, dpi)
    page[0].save(png_path, fmt)

def div(a, b, digit=4):
    if b == 0:
        return '-'
    else:
        return round(a / b, digit)

def make_BoP_sheet(total, excel_path, pdf_path, total_columns):
    wb = xl.load_workbook(excel_path)
    total_columns = [i for i in total_columns if '複勝' not in i]
    df = pd.DataFrame(columns = total_columns)
    sheetnames = list(wb.sheetnames)
    tansho_all_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}
    sanren_all_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}
    sheetnames = [i for i in sheetnames if i not in [total, total[0]+'月的中率', 'Sheet1']]
    # 回収金額の計算
    for name in sheetnames:
        ws_ = wb[name]
        # rankとvenue_idごとのbet数
        tansho_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}
        sanren_rank = {'A': 0, 'B': 0, 'C': 0, '-': 0, 'x': 0}  
        for i in range(2, 14):
            tansho_rank[ws_.cell(row=i, column=2).value] += 100
            sanren_rank[ws_.cell(row=i, column=2).value] += 2000
        for key in tansho_rank.keys():
            # tansho_rank及びsanren_rankを足し合わせたもの
            tansho_all_rank[key] += tansho_rank[key]
            sanren_all_rank[key] += sanren_rank[key]
        s = pd.Series([
            ws_['G16'].value,
            ws_['H16'].value,
            ws_['I16'].value,
            ws_['G17'].value,
            ws_['H17'].value,
            ws_['I17'].value,
            ws_['G18'].value,
            ws_['H18'].value,
            ws_['I18'].value,
            ws_['G19'].value,
            ws_['H19'].value,
            ws_['I19'].value,
            ws_['G20'].value,
            ws_['H20'].value,
            ws_['I20'].value
            ], name=name, index=total_columns).replace('-', np.nan)
        s = s-1
        for col in total_columns:
            if col[0] == '単' and 'ABC' in col:
                s[col] *= sum(tansho_rank.values()) - tansho_rank['-'] - tansho_rank['x']
            elif col[0] == '単' and col[-1] in ['A', 'B', 'C']:
                s[col] *= tansho_rank[col[-1]]
            elif col[0] == '単' and col[-1] not in ['A', 'B', 'C']:
                s[col] *= sum(tansho_rank.values()) - tansho_rank['x']
            if col[0] == '三' and 'ABC' in col:
                s[col] *= sum(sanren_rank.values()) - sanren_rank['-'] - sanren_rank['x']
            elif col[0] == '三' and col[-1] in ['A', 'B', 'C']:
                s[col] *= sanren_rank[col[-1]]
            elif col[0] == '三' and col[-1] not in ['A', 'B', 'C']:
                s[col] *= sum(sanren_rank.values()) - sanren_rank['x']
        df = df.append(s)
    total_s = pd.Series(df.sum(), name='合計(円)')
    total_s2 = pd.Series(index=total_columns, name='合計(%)', dtype=float)
    # 回収率の計算
    for col in total_columns:
        if col[0] == '単' and 'ABC' in col:
            total_s2[col] = div(total_s[col]+sum(tansho_all_rank.values()) - tansho_all_rank['-'] - tansho_all_rank['x'], sum(tansho_all_rank.values()) - tansho_all_rank['-'] - tansho_all_rank['x'])
        elif col[0] == '単' and col[-1] in ['A', 'B', 'C']:
            total_s2[col] = div(total_s[col]+tansho_all_rank[col[-1]], tansho_all_rank[col[-1]])
        elif col[0] == '単' and col[-1] not in ['A', 'B', 'C']:
            total_s2[col] = div(total_s[col]+sum(tansho_all_rank.values()) - tansho_all_rank['x'], sum(tansho_all_rank.values()) - tansho_all_rank['x'])
        if col[0] == '三' and 'ABC' in col:
            total_s2[col] = div(total_s[col]+sum(sanren_all_rank.values()) - sanren_all_rank['-'] - sanren_all_rank['x'], sum(sanren_all_rank.values()) - sanren_all_rank['-'] - sanren_all_rank['x'])
        elif col[0] == '三' and col[-1] in ['A', 'B', 'C']:
            total_s2[col] = div(total_s[col]+sanren_all_rank[col[-1]], sanren_all_rank[col[-1]])
        elif col[0] == '三' and col[-1] not in ['A', 'B', 'C']:
            total_s2[col] = div(total_s[col]+sum(sanren_all_rank.values()) - sanren_all_rank['x'], sum(sanren_all_rank.values()) - sanren_all_rank['x'])
    df = df.append(total_s)
    df = df.round(-1)
    df = df.append(total_s2)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    writer.book = wb
    writer.sheets = {ws_.title: ws_ for ws_ in wb.worksheets}
    df.to_excel(writer, sheet_name=total)
    writer.save()
    # 必要箇所にカラーリング
    ws = wb[total]
    side1 = Side(style='thin', color='000000')
    border1 = Border(top=side1, bottom=side1, left=side1, right=side1)
    for col in ws.columns:
        for i, cell in enumerate(col):
            coord = cell.coordinate
            ws[coord].border = border1
            cell.alignment = Alignment(horizontal = 'center', 
                                        vertical = 'center',
                                        wrap_text = False)
            if coord[0] == 'A' or coord[1:] == '1':
                ws[coord].fill = PatternFill(patternType='solid', fgColor='000000')
                ws[coord].font = Font(color='ffffff')
            else:
                ws[coord].fill = PatternFill(patternType='solid', fgColor='ffffff')
                ws[coord].font = Font(color='000000')
            try:
                if coord[1:] != str(ws.max_row):
                    if cell.value > 0:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='ffbf7f')
                    elif cell.value < 0:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='a8d3ff')
                else:
                    if cell.value > 1:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='ffbf7f')
                    elif cell.value < 1:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='a8d3ff')
                    cell.number_format = '0.00%'
            except:
                continue
        ws.column_dimensions[col[0].column_letter].width = 15
    # 月収支シートをファイルの先頭に移動させる
    for ws_ in wb.worksheets:
        i = i+1
        if total in ws.title:
            #先頭までの移動シート枚数
            Top_Sheet = i - 1
            #シート移動
            wb.move_sheet(ws,offset=-Top_Sheet)
    # xlsxをpdfに変換し、pdfディレクトリに保存する
    xlsx2pdf(pdf_path, ws, pagesize=[500, 250])
    # pdfをpngに変換し、pngディレクトリに保存する
    pdf2png(pdf_path)
    wb.save(excel_path)
    wb.close()

def make_hit_sheet(total, excel_path, pdf_path, total_columns):
    wb = xl.load_workbook(excel_path)
    total_columns = total_columns+['複勝(◎)', '複勝(○)', '複勝(▲)', 'ワイド', '三連複', '三連単']
    df = pd.DataFrame(columns = total_columns)
    sheetnames = list(wb.sheetnames)
    sheetnames = [i for i in sheetnames if i not in [total, total[0]+'月収支', 'Sheet1']]
    # 的中率の計算
    for name in sheetnames:
        ws_ = wb[name]
        s = pd.Series([
            ws_['C16'].value,
            ws_['D16'].value,
            ws_['E16'].value,
            ws_['F16'].value,
            ws_['C17'].value,
            ws_['D17'].value,
            ws_['E17'].value,
            ws_['F17'].value,
            ws_['C18'].value,
            ws_['D18'].value,
            ws_['E18'].value,
            ws_['F18'].value,
            ws_['C19'].value,
            ws_['D19'].value,
            ws_['E19'].value,
            ws_['F19'].value,
            ws_['C20'].value,
            ws_['E20'].value,
            ws_['F20'].value,
            ws_['B23'].value,
            ws_['C23'].value,
            ws_['D23'].value,
            ws_['E23'].value,
            ws_['F23'].value,
            ws_['G23'].value,
            ], name=name, index=total_columns).replace('0/0', np.nan)
        df = df.append(s)
    total_s = hit_sum(df, total_columns)
    df = df.append(total_s)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    writer.book = wb
    writer.sheets = {ws_.title: ws_ for ws_ in wb.worksheets}
    df.to_excel(writer, sheet_name=total)
    writer.save()
    # 必要箇所にカラーリング
    ws = wb[total]
    side1 = Side(style='thin', color='000000')
    border1 = Border(top=side1, bottom=side1, left=side1, right=side1)
    for col in ws.columns:
        for i, cell in enumerate(col):
            coord = cell.coordinate
            ws[coord].border = border1
            cell.alignment = Alignment(horizontal = 'center', 
                                        vertical = 'center',
                                        wrap_text = False)
            if coord[0] == 'A' or coord[1:] == '1':
                ws[coord].fill = PatternFill(patternType='solid', fgColor='000000')
                ws[coord].font = Font(color='ffffff')
            else:
                ws[coord].fill = PatternFill(patternType='solid', fgColor='ffffff')
                ws[coord].font = Font(color='000000')
            try:
                if coord[1:] != str(ws.max_row):
                    if cell.value > 0.5:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='ffbf7f')
                    elif cell.value < 0.5:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='a8d3ff')
                else:
                    if cell.value > 0.5:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='ffbf7f')
                    elif cell.value < 0.5:
                        ws[coord].fill = PatternFill(patternType='solid', fgColor='a8d3ff')
                    cell.number_format = '0.00%'
            except:
                continue
        ws.column_dimensions[col[0].column_letter].width = 15
    # 月的中率シートをファイルの先頭に移動させる
    for ws_ in wb.worksheets:
        i = i+1
        if total in ws.title:

            #先頭までの移動シート枚数
            Top_Sheet = i - 1
            #シート移動
            wb.move_sheet(ws,offset=-Top_Sheet)
    # xlsxをpdfに変換し、pdfディレクトリに保存する
    xlsx2pdf(pdf_path, ws, pagesize=[900, 250])
    # pdfをpngに変換し、pngディレクトリに保存する
    pdf2png(pdf_path)
    wb.save(excel_path)
    wb.close()

def hit_sum(df, total_columns):
    total_s = pd.Series(index=total_columns, name='合計(%)', dtype=float)
    for col in total_columns:
        total = df.loc[:, col]
        all_top = 0
        all_bottom = 0
        for a_hit in total:
            if isinstance(a_hit, str): top, bottom = [int(n) for n in a_hit.split('/')]
            else: top, bottom = (0, 0)
            all_top += top
            all_bottom += bottom
            total_s[col] = div(all_top, all_bottom)
    return total_s

def isExist_date(session, date):
    url = 'https://db.netkeiba.com/race/list/' + date
    html = session.get(url)
    html.encoding = "EUC-JP"
    soup = BeautifulSoup(html.content, "html.parser")
    # レースのある日付だけを引っ張ってこれる
    for true_date in soup.find_all('table')[0].find_all('td'):
        try:
            if re.findall(r'\d+', true_date.find('a').get('href'))[0] == date: return True
        except:
            pass
    return False

def make_venue_id_list(date, num='all'):
    venue_id_list = set()
    url = 'https://race.netkeiba.com/top/race_list_sub.html?kaisai_date=' + date
    r = requests.get(url)
    soup = BeautifulSoup(r.content, "html.parser")
    for race_list in soup.find_all('li'):
        venue_id_list.add(re.findall(r'\d+', race_list.find_all('a')[0].get('href'))[0][:10])
    if num=='all': return list(venue_id_list)
    else: return list(venue_id_list)[:num]

def make_race_class(shutuba_table, proba):
    df = shutuba_table.data_c.loc[proba[0], ['class_未勝利', 'class_１勝クラス', 'class_２勝クラス', 'class_３勝クラス', 'class_オープン']]
    for col in df.columns:
        if sum(df[col]):
            return col.strip('class_')
    return '新馬'