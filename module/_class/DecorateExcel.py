import openpyxl as xl
import numpy as np
import pandas as pd

from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side

from matplotlib.colors import rgb2hex
import matplotlib.pyplot as plt

class DecorateExcel:
    def __init__(self, df, excel_path, sheet_name, mode='append'):
        self.excel_path = excel_path # xlsxファイルの保存先
        self.sheet_name = sheet_name # 扱うsheet名
        self.row_heading = 1 # 行見出し
        self.col_heading = 1 # 列見出し
        self.row_length = 0 # 表の行の長さ
        self.col_length = 0 # 表の列の長さ
        self.mode = mode
        self.side = Side(style='thin', color='000000')
        self.border = Border(top=self.side, bottom=self.side, left=self.side, right=self.side)
        self.df = df
        self.df2xlsx()
    
    # wbを保存する関数
    def save(self):
        self.wb.save(self.excel_path)
        self.wb.close()
    
    # dfをxlsxに書き込む関数
    def df2xlsx(self):
        if self.mode == 'append':
            self.wb = xl.load_workbook(self.excel_path) # workbook
            writer = pd.ExcelWriter(self.excel_path, engine='openpyxl')
            writer.book = self.wb
            writer.sheets = {ws.title: ws for ws in self.wb.worksheets}
            startrow = writer.sheets[self.sheet_name].max_row+1
            self.df.to_excel(writer, sheet_name=self.sheet_name, startrow=startrow, startcol=0)
            writer.save()
        else:
            writer = pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
            self.df.to_excel(writer, sheet_name=self.sheet_name, startrow=0, startcol=0)
            writer.save()
            self.wb = xl.load_workbook(self.excel_path) # workbook
        self.ws = self.wb[self.sheet_name]
        self.row_length, self.col_length = self.df.shape
        self.col_heading = writer.sheets[self.sheet_name].max_row - self.row_length

    # 表の左上に表題を記述する関数
    def write_title(self, name = '', font = Font(bold=True, color='ffffff')):
        self.ws.cell(column = self.row_heading, row = self.col_heading).value = name
        self.ws.cell(column = self.row_heading, row = self.col_heading).font = font

    # cellが表の見出しかどうか判定する関数
    def is_heading_cell(self, cell):
        return (cell.column == self.row_heading) or (cell.row == self.col_heading)
    
    # あるレースの本命馬の着順と入力した着順が一致しているかどうか判定する関数
    def is_hit_order(self, cell, order):
        return '('+str(order)+')' in self.ws.cell(column=self.col_index('1着予想◎'), row=cell.row).value

    # あるレースの三連複が的中したかどうか判定する関数
    def is_hit_sanrenpuku(self, cell, num):
        race_data = self.df.loc[self.ws.cell(column = self.row_heading, row = cell.row).value]
        race_data_index = race_data.index.values
        cnt = 0
        for data, index in zip(race_data, race_data_index):
            if '予想' not in index: continue
            for order in ['(1)', '(2)', '(3)']:
                cnt += order in data
        return cnt == num
    
    # あるレースの三連単が的中したかどうか判定する関数
    def is_hit_sanrentan(self, cell, num):
        return self.is_hit_sanrenpuku(cell, num) and self.is_hit_order(cell, 1)

    # 入力した名前の列名のindexを返す関数
    def col_index(self, name):
        for i in range(self.col_length):
            if self.ws.cell(column=i+1, row=self.col_heading).value == name:
                return i+1
        raise '{} does not exist in the column'.format(name)

    # 入力したcellの列名を返す関数
    def col_name(self, cell):
        return self.ws.cell(column=cell.column, row=self.col_heading).value

    # 対象の列名を持つcellのフォーマットを%に変更する関数
    def to_percent(self, cell, pct_list):
        # セルのフォーマットをパーセンテージにする
        if self.col_name(cell) in pct_list:
            cell.number_format = '0.00%'
    
    # 馬のスコアごとにカラーリング(赤色のグラデーション)
    def coloring_by_score(self, proba_table, venue_id):
        cmap = plt.get_cmap('Reds')
        for row in self.ws.rows:
            if self.col_heading < row[0].row <= self.col_heading + self.row_length:
                for cell, score in zip(row[3:9], proba_table.loc[venue_id+str(row[0].row-1).zfill(2)].sort_values('score', ascending = False)['score'].head(6)):
                    coord = cell.coordinate
                    score = max(min(score, 3.5), 0)
                    colorcode = rgb2hex(cmap(score/3.5)).replace('#', '')
                    self.ws[coord].fill = PatternFill(patternType='solid', fgColor=colorcode)

    # xlsxファイルに色付けなど装飾する関数
    def decorate_excelsheet(
                            self,
                            mapping_dict={
                                            'ffbf7f': {'rank': 'A', 'order': 1},
                                            'a8d3ff': {'rank': 'B', 'order': 2},
                                            'd3d3d3': {'rank': 'C', 'order': 3},
                                            '50C878': {'rank': 'ABC', 'order': 0}
                                        },
                            pct_list=[],
                            heading_color='000000'
                            ):
        for col in self.ws.columns:
            max_length = 0
            for cell in col:
                coord = cell.coordinate
                max_length = max(max_length, len(str(cell.value)))
                # 文字を中心に配置する
                cell.alignment = Alignment(
                                            horizontal = 'center', 
                                            vertical = 'center',
                                            wrap_text = False
                                            )
                # 範囲外のセルへの変更は施さない
                if cell.row < self.col_heading-1 or cell.row > self.col_heading + self.row_length: continue
                # 表の間と右の余白は白色にする
                elif cell.row == self.col_heading-1 or cell.column > self.col_length+1:
                    self.ws[coord].fill = PatternFill(patternType='solid', fgColor='ffffff')
                    self.ws[coord].font = Font(color='000000')
                    continue
                # 行見出し、列見出しは黒色にする
                elif self.is_heading_cell(cell):
                    self.ws[coord].fill = PatternFill(patternType='solid', fgColor=heading_color)
                    self.ws[coord].font = Font(color='ffffff')
                    # 特定の値を持った行見出し、列見出しのセルの色を変更する
                    for color, value_dict in mapping_dict.items():
                        if self.ws[coord].value == value_dict['rank']:
                            self.ws[coord].fill = PatternFill(patternType='solid', fgColor=color)
                            self.ws[coord].font = Font(color='000000')
                else:
                    self.ws[coord].fill = PatternFill(patternType='solid', fgColor='ffffff')
                    self.ws[coord].font = Font(color='000000')
                    self.to_percent(cell, pct_list)
                    # 特定の条件を満たした、見出し以外のセルの色を変更する
                    for color, value_dict in mapping_dict.items():
                        if (    self.ws[coord].value == value_dict['rank']
                            or (self.col_name(cell) == '単勝オッズ' and self.is_hit_order(cell, value_dict['order']))
                            or (self.col_name(cell) == '三連複オッズ' and self.is_hit_sanrenpuku(cell, 4-value_dict['order']))
                            or (self.col_name(cell) == '三連単オッズ' and self.is_hit_sanrentan(cell, 4-value_dict['order']))):
                            self.ws[coord].fill = PatternFill(patternType='solid', fgColor=color)
                            self.ws[coord].font = Font(color='000000')
                self.ws[coord].border = self.border
            adjusted_width = max_length * 2.08
            self.ws.column_dimensions[cell.column_letter].width = adjusted_width
        self.write_title(self.sheet_name)